from tkinter import *
import tkinter.ttk as ttk
import tkinter.font
from tkinter import filedialog
import re
import os
import time
from datetime import datetime
import math
from collections import Counter
from openpyxl import Workbook
from openpyxl.styles import Alignment, Font, Border, Side, PatternFill
import sys
import time
import getpass
import pdb


dCount = {}
dList = []
wb = Workbook()
ws = wb.worksheets[0]
ws.title = 'Training data'



root = Tk()
root.title("Lionbridge Bixby Training Analyzer v.1.0")
# root.iconbitmap(default = 'monitoring.ico')

def openTom():
    path = "./"
    path = os.path.realpath(path)
    os.startfile(path)


def add_folder():

    outpath = filedialog.askdirectory(initialdir="/", title="폴더를 선택하세요")


    print(outpath)
    outpath = outpath+"/"

    # outpath = "C:/reminder/reminderApp/resources/ko-KR/training/"
    # outpath = "C:/Users/LBuser/PycharmProjects/LB_TRAINER/venv/testmonial/"

    now = datetime.now().strftime('%Y-%m-%d_%H-%M-%S')
    file_list = os.listdir(outpath)
    file_list.sort()

    utt_kr = []
    start_time = time.time()

    # 파일들을 하나씩 다 뽑기
    for n in file_list:

        file_path = outpath + n
        statusWindow.insert(END, "\n 파일 로드 중 : " + file_path)
        print("파일 로드 중 : " + file_path)
        f = open(file_path, 'r', encoding='UTF8')

        while True:
            line = f.readline()
            if not line: break
            if 'utterance' in line:
                # pure_nl은 nl 그 자체.
                pure_nl = line[13:-2].replace('"', '')
                if pure_nl.startswith(" "): pure_nl = pure_nl[1:]
                utt_kr.append(pure_nl)

    statusWindow.insert(END, "\n파일 로드 완료.")
    print("파일 로드 완료.")
    utt_kr.sort()  # nl([g:][v:]값) 전체 정렬
    cnt = 0
    # pdb.set_trace()

    statusWindow.insert(END, "\n골 오름차순 순서 재 배열 완료.")
    print("골 오름차순 순서 재 배열 완료.")

    goal_list = []
    impurity = []
    impurity_nl = []

    goalKeeper = []
    pure_krKeeper = []
    sortaKeeper = []

    # 데이터를 순서대로 재배열 후 엑셀로 저장
    for sorta in utt_kr:
        cnt += 1

        # x.write("\n\n" + n + "\n\n")

        if sorta.startswith(" "): sorta = sorta[1:]  # 앞에 공백으로 시작하면 공백 지워버리기

        goal_index = sorta.find("]")
        goal = sorta[3:goal_index]
        # x.write(str(cnt) + ". " + sorta + "\n" + goal+ "\n")
        goal_list.append(goal)

        # pure_kr은 학습된 그 자체 한글 내용 텍스트.
        pure_kr = re.sub('\[[^)]*\]', '', sorta)
        pure_kr = re.sub('[(|)|{|}]', '', pure_kr)
        if pure_kr.startswith(" "): pure_kr = pure_kr[1:]

        # x.write(pure_kr + "\n")

        # Excel 저장

        routeDesktop = "C:/"+ getpass.getuser()+"/Desktop"

        goalKeeper.append(goal)
        pure_krKeeper.append(pure_kr)
        sortaKeeper.append(sorta)

        # ws['A' + str(cnt)] = goal
        # ws['B' + str(cnt)] = pure_kr
        # ws['C' + str(cnt)] = sorta

        hangul = re.compile('[^ ㄱ-ㅣ가-힣|0-9]+')  # 한글과 숫자 띄어쓰기를 제외한 모든 글자
        result = hangul.sub('', pure_kr)  # 한글과 띄어쓰기를 제외한 모든 부분을 제거
        # print(result)
        #
        # print("sorta : ", sorta)
        dList.append(sorta)

        if hangul.findall(pure_kr):
            impurity.append(sorta)
            impurity_nl.append(hangul.findall(pure_kr))


        # print(hangul.findall(pure_kr))
        # routeDesktop = "/home/"+ getpass.getuser()+"/Desktop/"
        #
        #
        # print(routeDesktop)
        #
        # currentPath = os.getcwd()
        #
        # print(currentPath)
        # os.chdir(routeDesktop)
        # print(os.getcwd())




        statusWindow.insert(END, "\n" + sorta)
        # print(goal + " : " + pure_kr + " / " + sorta)


        proceed = (int(cnt) / int(len(utt_kr))) * 100
        proceed = math.floor(proceed)


        statusWindow.insert(END, "\n진행중 .. " + str(proceed) + " %")
        statusWindow.see(END)
        print("진행중 .. " + str(proceed) + " %")

        # print("route :" + route)

        p_var2.set(proceed)
        progressbar.update()
        print(p_var2.get())
    

    for i in range(0,len(goalKeeper)):

        ws.cell(row=i+1, column=1).value = goalKeeper[i]
        ws.cell(row=i+1, column=2).value = pure_krKeeper[i]
        ws.cell(row=i+1, column=3).value = sortaKeeper[i]


    end_time = time.time()

    elapsed = end_time - start_time
    elapsed = math.floor(elapsed)
    statusWindow.insert(END, "\n.\n.\n.\n.소요시간 : " + str(elapsed) + " 초")
    statusWindow.insert(END, "\n총 학습발화 개수 : " + str(cnt))
    statusWindow.see(END)



######여기서부터는 별도의 윈도우로.

    # 골별 개수 파악
    print("소요시간 : " + str(elapsed) + " 초")
    resultWindow.insert(END, "\n소요시간 : " + str(elapsed) + " 초")
    resultWindow.see(END)

    # 총 발화개수
    print("총 학습발화 개수 : " + str(cnt) +"개")
    resultWindow.insert(END, "\n총 학습발화 개수 : " + str(cnt) +"개")
    resultWindow.see(END)


    collect_goals = Counter(goal_list)
    print(collect_goals)

    collect_goals_sorted = sorted(collect_goals.items(), reverse=True, key=lambda item: item[1])
    print(collect_goals_sorted)

    print("골 별 발화 수 : ")
    resultWindow.insert(END, "\n\n[골 별 발화 수]\n")


    way1 = []

    for k, v in collect_goals_sorted:
        print(k, ":", v)
        txt1 = "\n" + str(k) + "  :  "+ str(v)
        way1.append(txt1)

    for aa in way1:
        resultWindow.insert(END, str(aa))

    print("완료.")


    print(goal_list)

    # 불순물 제거
    print("불순물 발화 :")


    impu_cnt = 0

    way2 = []

    for d in impurity:
        print(str(impurity_nl[impu_cnt]), str(d))
        txt2 = "\n\n[" + str(impurity_nl[impu_cnt]) + " 발견]\n\n" + str(d)
        way2.append(txt2)

        impu_cnt += 1
    impuWindow.insert(END, "\n [오류검출 대표 유형 정리]\n")
    impuWindow.insert(END, "\n [" + r'\n' + "] : 학습데이터에 엔터키값이 포함됨")
    impuWindow.insert(END, "\n [" + r'\xa0' + "] : 학습데이터에 공백 유니코드값이 포함됨")
    impuWindow.insert(END, "\n [�] : 학습데이터에 대치문자(�) 포함됨. 주로 IDE 오류로 발생.")
    impuWindow.insert(END, "\n\n 위의 형태가 발견되면 Bixby IDE에서 직접 수정바랍니다.\n")
    impuWindow.insert(END, "\n----------------------------------------------------------\n\n\n")

    for bb in way2:
        impuWindow.insert(END, str(bb))




    # 중복값 검출
    # print("중복값 체크를 위한 리스트 - dList : \n", dList)

    for i in dList:
        try:
            dCount[i] += 1

        except:
            dCount[i] = 1

    way3 = []
    way3_cnt = 0
    for xx in dCount:
        if dCount[xx] > 1:
            txt3 = "\n\n[=====" + str(dCount[xx]) +"개 발견=====]\n" + str(xx)
            # print(txt)
            way3.append(txt3)
            way3_cnt += 1

    for cc in way3:
        duplicatedWindow.insert(END, str(cc))

    statusWindow.insert(END, "\n\n예상 오탈자 발견 : " + str(impu_cnt))
    statusWindow.insert(END, "\n\n중복 발화 검출 개수: " + str(way3_cnt))

    statusWindow.see(END)



    #### 분석결과 요약시트 작성

    ws2 = wb.create_sheet('분석 결과')
    font_bold = Font(size=12, bold=True, color='000000')  # 000000: black

    ws2['B2'] = "ko-KR Training Data 분석결과 "
    ws2['B2'].font = font_bold
    ws2['B3'] = datetime.now().strftime("%Y-%m-%d %H:%M")

    ws2['B6'] = "소요시간 : "
    ws2['C6'] = str(elapsed) + " 초"
    ws2['B7'] = "총 학습발화 개수 : "
    ws2['C7'] = str(cnt)+ " 개"
    ws2['B8'] = "예상 오탈자 발견 : "
    ws2['C8'] = str(impu_cnt) + " 개"
    ws2['B9'] = "중복 발화 검출 개수: "
    ws2['C9'] = str(way3_cnt) + " 개"

    ws2['F2'] = "[골 별 발화 수]"

    cell_index = 3
    for aa in way1:
        ws2['F'+str(cell_index)] = str(aa)
        print(aa)
        cell_index += 1
    ws2.column_dimensions['B'].width = 20
    ws2.column_dimensions['F'].width = 65



    #### 오탈자 리스트 엑셀저장

    ws3 = wb.create_sheet('오탈자 결과')
    font_bold = Font(size=12, bold=True, color='000000')  # 000000: black

    ws3['B2'] = "오탈자 결과 "
    ws3['B2'].font = font_bold

    ws3['B4'] = "예상 오탈자 발견 : "
    ws3['C4'] = str(impu_cnt) + " 개"

    b_cell_index = 2
    for bb in way2:
        ws3['F' + str(b_cell_index)] = str(bb)
        print(bb)
        b_cell_index += 1

    ws3.column_dimensions['B'].width = 20
    ws3.column_dimensions['F'].width = 65


    #### 중복 발화 리스트 엑셀저장

    ws4 = wb.create_sheet('중복 발화 결과')
    font_bold = Font(size=12, bold=True, color='000000')  # 000000: black

    ws4['B2'] = "중복발화 결과 "
    ws4['B2'].font = font_bold

    ws4['B4'] = "중복발화 발견 : "
    ws4['C4'] = str(way3_cnt) + " 개"

    print('way3 디버그')
    print(way3)
    c_cell_index = 2
    for cc in way3:
        ws4['F' + str(c_cell_index)] = str(cc)
        print(cc)
        c_cell_index += 1

    ws4.column_dimensions['B'].width = 20
    ws4.column_dimensions['F'].width = 65





    # sheet_ws = ws.active
    # sheet_ws2.cell(row=1, column=3).value = '소요시간'
    # # sheet_ws2.cell(row=1, column=4).value = pure_krKeeper[i]
    # # sheet_ws2.cell(row=1, column=5).value = sortaKeeper[i]

    wb.save("rawdata_" + str(now) + ".xlsx")
    wb.close()

# GUI

root.geometry("1624x600")
root.resizable(False, False)

font = tkinter.font.Font(size=14)


#Frame1

frame1 = Frame(root, relief="solid", bd=1)
frame1.grid(row=0, column=0)

title = Label(frame1, text="Lionbridge Bixby Training Analyzer\nv.1.0.210907\n\nv-euisung.lee", foreground="#000000", font=font, width=50, height=5)
title.grid(row=0, column=0)

# blank = Label(frame1, font=font, width=7, height=3)
# blank.grid(row=0, column=1)

btnFileLoad = Button(frame1, text="시작하려면 여기를 누르세요.\n(training 파일이 들어간 폴더를 선택해주세요.)", font=font, command=add_folder, background="#CC2A49",foreground="#FFFFFF",height=5)
btnFileLoad.grid(row=1, column=0, sticky='news')


p_var2 = DoubleVar()
progressbar = ttk.Progressbar(frame1, maximum=100, variable=p_var2)
progressbar.grid(row=2, column=0, sticky='news')




statusScroll = Scrollbar(frame1)
statusWindow = Text(frame1, width=69, yscrollcommand=statusScroll.set)

statusScroll.config(command=statusWindow.yview)

statusWindow.grid(row=3, column=0, sticky='nw')
statusScroll.grid(row=3, column=0, sticky='nse')


btnExcel = Button(frame1, text="분류된 엑셀 보기", font=font, command=openTom, background="#F99E4C",foreground="#FFFFFF",height=2)
btnExcel.grid(row=4, column=0, sticky='news')

now = datetime.now().strftime('%Y-%m-%d  %H:%M')

statusWindow.insert(END, "\n\n  분석이 완료되면 엑셀파일이 자동으로 생성되며\n\n  아래 '분류된 엑셀보기'를 누르면\n\n  엑셀파일이 저장된 폴더가 열립니다.\n")
statusWindow.insert(END, "\n\n------------------------------------------------\n")
statusWindow.insert(END, "\n\n  " + str(now))
statusWindow.insert(END, "\n\n  분석 대기중..")



#Frame2
frame2 = Frame(root, relief="solid", bd=1)
frame2.grid(row=0, column=1, sticky='news')

result = Label(frame2, text="골 별 발화 개수", background="#F36F38", foreground="#FFFFFF", font=font, width=50, height=5)
result.grid(row=0, column=0, sticky='news')


resultScroll = Scrollbar(frame2)
resultWindow = Text(frame2, width=69, height=38,yscrollcommand=resultScroll.set)

resultScroll.config(command=resultWindow.yview)

resultWindow.grid(row=1, column=0, sticky='nsw')
resultScroll.grid(row=1, column=0, sticky='nse')






#Frame3
frame3 = Frame(root)
frame3.grid(row=0, column=2, sticky='news')

impu = Label(frame3, text="발화 오탈자 필터 (한글/숫자를 제외한 영문&특수문자 포함발화)", background="#EF4648", foreground="#FFFFFF", font=font, width=60, height=2)
impu.grid(row=0, column=0, sticky='news')

impuScroll = Scrollbar(frame3)
impuWindow = Text(frame3, width=84, height=19, yscrollcommand=impuScroll.set)

impuScroll.config(command=impuWindow.yview)

impuWindow.grid(row=1, column=0, sticky='nsw')
impuScroll.grid(row=1, column=0, sticky='nse')



dupli = Label(frame3, text="중복발화 검출 결과", background="#582841", foreground="#FFFFFF", font=font, height=2)
dupli.grid(row=2, column=0, sticky='news')

duplicatedScroll = Scrollbar(frame3)
duplicatedWindow = Text(frame3, width=84, height=20, yscrollcommand=duplicatedScroll.set)

duplicatedScroll.config(command=duplicatedWindow.yview)

duplicatedWindow.grid(row=3, column=0, sticky='nsw')
duplicatedScroll.grid(row=3, column=0, sticky='nse')

root.mainloop()
